"""
=============================================================================
SCRIPT NAME: build_gdelt_panel.py
=============================================================================

INPUT FILES:
- /Users/arjundivecha/Dropbox/AAA Backup/A Working/GDELT/Deep/data/features/country_signal_monthly_deep.parquet
    Single source-of-truth monthly panel — already merges shallow GDELT
    signals (metronome / risk / sentiment / attention / tone families) with
    the deep ingest (themes, GCAM, events). Produced by the Deep pipeline
    at /A Working/GDELT/Deep/.

OUTPUT FILES:
- GDELT.xlsx (this directory)
    Wide monthly workbook consumed by:
      Step Two GDELT Create Tidy.py
    Layout matches what export_deep_workbook.py produces today:
      Sheet 1: README        — architecture overview, feature families
      Sheet 2: README_VARIABLES — variable dictionary
      Sheets 3..N: one per variable, rows = month-end dates,
                   columns = 34 country buckets in canonical order.
    Step Two GDELT skips README / README_VARIABLES, normalizes month-end
    → first-of-month via pandas to_period("M").to_timestamp().

VERSION: 1.0
LAST UPDATED: 2026-04-28
AUTHOR: Arjun Divecha (with Claude)

DESCRIPTION:
Generates GDELT.xlsx — the input panel for the Pure GDELT track of the
T2 GDELT pipeline — directly in this directory, with atomic write and
timestamped backup. Replaces the old workflow where:
  1. /A Complete/GDELT/scripts/export_deep_workbook.py was run manually
  2. The resulting xlsx was manually copied to /A Complete/T2 GDELT/

This is an INTERIM step. The longer-term plan (per the ASADO consolidation
roadmap) is to move both the shallow and deep GDELT pipelines, plus this
exporter, into the ASADO repo. For now this script just stops the manual
copy/rename and wraps the existing exporter with safety guards.

The cycle protection: this script reads only the upstream Deep parquet,
NOT from any DuckDB view that could ever include optimizer outputs. Same
rule as build_econ_panel.py.

DEPENDENCIES:
- pandas, openpyxl, pyarrow
- /A Complete/GDELT/scripts/export_deep_workbook.py (delegate)

USAGE:
  python3 build_gdelt_panel.py
  python3 build_gdelt_panel.py --dry-run     # report counts, do not write
  python3 build_gdelt_panel.py --panel-parquet <other-path>
=============================================================================
"""

from __future__ import annotations

import argparse
import logging
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_PANEL_PARQUET = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/A Working/GDELT/Deep/"
    "data/features/country_signal_monthly_deep_treated.parquet"
)
EXPORTER_SCRIPT = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/A Complete/GDELT/scripts/"
    "export_deep_workbook.py"
)
OUTPUT_PATH = SCRIPT_DIR / "GDELT.xlsx"
BACKUP_DIR = SCRIPT_DIR / "backups"
LOCK_FILE = SCRIPT_DIR / "~$GDELT.xlsx"

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


def warn_if_locked() -> None:
    if LOCK_FILE.exists():
        logger.warning(
            "Excel lock file present (%s) — GDELT.xlsx may be open in Excel. "
            "Close it before running, or the rename will fail.", LOCK_FILE.name,
        )


def backup_existing(output_path: Path) -> Optional[Path]:
    if not output_path.exists():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y_%m_%d_%H%M%S")
    backup_path = BACKUP_DIR / f"{output_path.stem}_{timestamp}{output_path.suffix}"
    shutil.copy2(output_path, backup_path)
    logger.info("Backed up existing %s -> %s", output_path.name, backup_path)
    return backup_path


def run_exporter(panel_parquet: Path, tmp_output: Path) -> None:
    """Delegate to export_deep_workbook.py for the heavy lifting."""
    cmd = [
        sys.executable,
        str(EXPORTER_SCRIPT),
        "--panel-parquet", str(panel_parquet),
        "--output", str(tmp_output),
    ]
    logger.info("Running: %s", " ".join(cmd))
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.stdout.strip():
        for line in result.stdout.strip().splitlines():
            logger.info("[exporter] %s", line)
    if result.returncode != 0:
        if result.stderr.strip():
            for line in result.stderr.strip().splitlines():
                logger.error("[exporter] %s", line)
        raise RuntimeError(f"exporter failed with exit code {result.returncode}")


def verify_output(path: Path) -> int:
    """Sanity-check the produced workbook. Returns the sheet count."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_count = len(wb.sheetnames)
    docs = [s for s in wb.sheetnames if s in ("README", "README_VARIABLES")]
    wb.close()
    logger.info("Output sanity: %d sheets total (%d doc, %d data)",
                sheet_count, len(docs), sheet_count - len(docs))
    if sheet_count < 100:
        raise RuntimeError(
            f"Sheet count too low ({sheet_count}); expected ~1100+. "
            "Aborting before overwriting good copy."
        )
    return sheet_count


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build GDELT.xlsx panel for the T2 GDELT pipeline."
    )
    parser.add_argument("--panel-parquet", type=Path, default=DEFAULT_PANEL_PARQUET,
                        help="Source monthly panel parquet (shallow + deep merged)")
    parser.add_argument("--output", type=Path, default=OUTPUT_PATH,
                        help=f"Output xlsx (default: {OUTPUT_PATH.name} in this dir)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Skip the actual write; just verify the parquet exists")
    args = parser.parse_args()

    if not args.panel_parquet.exists():
        logger.error("Panel parquet not found: %s", args.panel_parquet)
        return 2
    if not EXPORTER_SCRIPT.exists():
        logger.error("Exporter not found: %s", EXPORTER_SCRIPT)
        return 2

    if args.dry_run:
        size_mb = args.panel_parquet.stat().st_size / 1e6
        logger.info("dry-run: would read %s (%.1f MB) and write %s",
                    args.panel_parquet, size_mb, args.output)
        return 0

    warn_if_locked()
    # Temp name keeps .xlsx so openpyxl can sanity-check it before rename.
    tmp_output = args.output.with_name(f".building.{args.output.name}")
    if tmp_output.exists():
        tmp_output.unlink()

    run_exporter(args.panel_parquet, tmp_output)
    verify_output(tmp_output)
    backup_existing(args.output)
    tmp_output.replace(args.output)
    size_mb = args.output.stat().st_size / 1e6
    logger.info("Wrote %s (%.1f MB)", args.output, size_mb)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
