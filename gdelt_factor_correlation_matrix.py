"""
GDELT Factor Correlation Matrix Generator

Creates a correlation matrix of factor returns grouped by category,
with mean return, standard deviation, and Sharpe ratio statistics.

INPUT FILES:
- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/T2 GDELT/GDELT_Optimizer.xlsx
  Monthly factor returns for 92 factors across 128 time periods (Date column + factor columns)

- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/T2 GDELT/Step Factor Categories GDELT.xlsx
  Mapping of 92 factor names to 16 categories (Factor Name, Category, Max columns)

OUTPUT FILES:
- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/T2 GDELT/GDELT_Factor_Correlation_Matrix.xlsx
  Excel file with:
  - Correlation matrix of all factors, columns/rows grouped by category
  - Mean return row below the matrix
  - Standard deviation row below mean
  - Sharpe ratio row at the bottom
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


def main():
    # =========================================================================
    # INPUT
    # =========================================================================
    input_dir = Path("/Users/arjundivecha/Dropbox/AAA Backup/A Complete/T2 GDELT")
    optimizer_file = input_dir / "GDELT_Optimizer.xlsx"
    categories_file = input_dir / "Step Factor Categories GDELT.xlsx"

    # =========================================================================
    # OUTPUT
    # =========================================================================
    output_file = input_dir / "GDELT_Factor_Correlation_Matrix.xlsx"

    # =========================================================================
    # READ DATA
    # =========================================================================
    optimizer_df = pd.read_excel(optimizer_file)
    categories_df = pd.read_excel(categories_file)

    # Set Date as index
    optimizer_df = optimizer_df.set_index("Date")

    # =========================================================================
    # BUILD FACTOR ORDER GROUPED BY CATEGORY
    # =========================================================================
    # Sort factors by category, then by factor name within each category
    categories_sorted = categories_df.sort_values(["Category", "Factor Name"]).reset_index(drop=True)

    # Filter to only factors that exist in the optimizer data
    available_factors = categories_sorted[categories_sorted["Factor Name"].isin(optimizer_df.columns)]
    ordered_factors = available_factors["Factor Name"].tolist()

    # Create a mapping from factor name to category
    factor_to_category = dict(zip(available_factors["Factor Name"], available_factors["Category"]))

    # Get the ordered list of unique categories
    ordered_categories = available_factors["Category"].unique().tolist()

    # Build category to factors mapping
    category_factors = {}
    for cat in ordered_categories:
        category_factors[cat] = available_factors[available_factors["Category"] == cat]["Factor Name"].tolist()

    print(f"Total factors in correlation matrix: {len(ordered_factors)}")
    print(f"Categories: {ordered_categories}")
    for cat in ordered_categories:
        print(f"  {cat}: {len(category_factors[cat])} factors")

    # =========================================================================
    # COMPUTE CORRELATION MATRIX
    # =========================================================================
    corr_matrix = optimizer_df[ordered_factors].corr()

    # =========================================================================
    # COMPUTE STATISTICS (monthly)
    # =========================================================================
    returns = optimizer_df[ordered_factors]

    # Mean monthly return
    mean_returns = returns.mean()

    # Monthly standard deviation
    std_returns = returns.std()

    # Sharpe ratio (monthly): mean / std
    # Annualized Sharpe = (mean / std) * sqrt(12)
    sharpe_ratios = (mean_returns / std_returns) * np.sqrt(12)

    # =========================================================================
    # BUILD OUTPUT DATAFRAME
    # =========================================================================
    # The output will have:
    # - Rows: factors (grouped by category) + statistics rows
    # - Columns: factors (grouped by category)

    n_factors = len(ordered_factors)

    # Create the full output matrix with correlation values
    output_data = corr_matrix.copy()

    # Add statistics rows at the bottom
    stats_labels = ["Mean Return", "Std Dev", "Sharpe Ratio (Ann.)"]

    for label in stats_labels:
        row_data = pd.Series(dtype=float)
        if label == "Mean Return":
            row_data = mean_returns
        elif label == "Std Dev":
            row_data = std_returns
        elif label == "Sharpe Ratio (Ann.)":
            row_data = sharpe_ratios

        output_data.loc[label] = row_data

    # =========================================================================
    # WRITE TO EXCEL WITH FORMATTING
    # =========================================================================
    writer = pd.ExcelWriter(output_file, engine="openpyxl")
    output_data.to_excel(writer, sheet_name="Correlation Matrix", startrow=2, startcol=1, index=True)

    wb = writer.book
    ws = wb["Correlation Matrix"]

    # ===== ADD TITLE AND CATEGORY HEADERS =====
    # Title
    title_cell = ws.cell(row=1, column=2, value="GDELT Factor Correlation Matrix (Grouped by Category)")
    title_cell.font = Font(bold=True, size=16)

    # Column headers: add category labels above factor columns
    col_idx = 2  # column B (first factor column)
    for cat in ordered_categories:
        factors_in_cat = category_factors[cat]
        start_col = col_idx
        end_col = col_idx + len(factors_in_cat) - 1

        # Write category name in the row above (row 2)
        cat_cell = ws.cell(row=2, column=start_col, value=cat)
        cat_cell.font = Font(bold=True, size=11, color="FFFFFF")
        cat_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cat_cell.alignment = Alignment(horizontal="center")

        if end_col > start_col:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)

        col_idx = end_col + 1

    # Row headers: add category labels to the left of factor rows
    # The data starts at row 3 (row 1 = title, row 2 = col headers, row 3 = first data row)
    row_idx = 3
    for cat in ordered_categories:
        factors_in_cat = category_factors[cat]
        start_row = row_idx
        end_row = row_idx + len(factors_in_cat) - 1

        # The factor names are in column B (index 2), so category label goes in column A (index 1)
        # Write category name
        cat_label_cell = ws.cell(row=start_row, column=1, value=cat)
        cat_label_cell.font = Font(bold=True, size=11, color="FFFFFF")
        cat_label_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cat_label_cell.alignment = Alignment(horizontal="center", vertical="center")

        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        row_idx = end_row + 1

    # ===== FORMAT CORRELATION VALUES =====
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Color scale for correlations: green (positive) to red (negative)
    # We'll manually color cells based on values
    data_start_row = 3
    data_end_row = data_start_row + n_factors - 1
    data_start_col = 2
    data_end_col = data_start_col + n_factors - 1

    for row in range(data_start_row, data_end_row + 1):
        for col in range(data_start_col, data_end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = "0.000"
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Color based on correlation value
            val = cell.value
            if val is not None and isinstance(val, (int, float)):
                if val > 0.7:
                    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif val > 0.4:
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                elif val > 0.2:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif val < -0.7:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif val < -0.4:
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif val < -0.2:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ===== FORMAT STATISTICS ROWS =====
    stats_row_start = data_end_row + 2  # one blank row separator

    # Add section label
    stats_label_cell = ws.cell(row=stats_row_start - 1, column=1, value="STATISTICS")
    stats_label_cell.font = Font(bold=True, size=12, color="2F5496")

    stats_labels_for_rows = ["Mean Return", "Std Dev", "Sharpe Ratio (Ann.)"]
    stats_fills = [
        PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    ]

    for i, label in enumerate(stats_labels_for_rows):
        row_num = stats_row_start + i
        # Format the label cell
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.fill = stats_fills[i]

        # Format value cells
        for col in range(data_start_col, data_end_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.number_format = "0.000"
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.fill = stats_fills[i]

            # Highlight significant values
            val = cell.value
            if val is not None and isinstance(val, (int, float)):
                if label == "Sharpe Ratio (Ann.)" and abs(val) > 1.0:
                    cell.font = Font(bold=True, color="006100")

    # ===== FORMAT FACTOR NAME CELLS =====
    # Row labels (column B)
    for row in range(data_start_row, data_end_row + 1):
        cell = ws.cell(row=row, column=2)
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="left")

    # Column labels (row 3)
    for col in range(data_start_col, data_end_col + 1):
        cell = ws.cell(row=data_start_row - 1, column=col)
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="center", text_rotation=90)

    # ===== SET COLUMN WIDTHS =====
    ws.column_dimensions["A"].width = 15  # Category column
    ws.column_dimensions["B"].width = 30  # Factor name row labels

    for col in range(data_start_col, data_end_col + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 8

    # ===== FREEZE PANES =====
    ws.freeze_panes = "C4"

    # ===== SET PRINT AREA AND PAGE SETUP =====
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ===== SAVE =====
    writer.close()

    print(f"\nOutput saved to: {output_file}")
    print(f"Correlation matrix: {n_factors}x{n_factors} factors")
    print(f"Categories: {len(ordered_categories)}")
    print(f"Statistics: Mean Return, Std Dev, Sharpe Ratio (Annualized)")


if __name__ == "__main__":
    main()
